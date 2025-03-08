# Author: 邵世昌
# CreatTime: 2024/11/23
# FileName: 修改excel内容
from openpyxl import load_workbook
workbook = load_workbook('test.xlsx')
sheet = workbook['Sheet1']
sheet.cell(row=1, column=2).value = '这是修改后的内容'
workbook.save('test.xlsx')
