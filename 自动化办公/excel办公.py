# Author: 邵世昌
# CreatTime: 2024/11/23
# FileName: excel办公
from openpyxl import load_workbook
# 打开excel文件
wb = load_workbook(r'C:\Users\25782\Desktop\test.xlsx')
# 选择页数
sheet = wb.worksheets[0]
# sheet = wb.get_sheet_by_name('sheet1') 方法2
# 选择操作内容
#%%
for row in sheet.iter_rows(min_row=1):
    print(row[1].value,row[2].value)

#%%
for row in sheet.iter_rows(min_col=1,max_col=3):
    for cell in row:
        print(cell.value)

#%% 获取所有数据
rows = []
for row in sheet.columns:
    rows.append([cell.value for cell in row])
print(rows)